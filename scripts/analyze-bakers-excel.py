#!/usr/bin/env python3
"""Analyze Bakers Mart Excel — print all sheets, columns, and sample data."""
import openpyxl

FILE = '/home/z/my-project/upload/Bakers_Mart_DMP_Business_Report_April_to_June_2026.xlsx'

# Load without read_only so we get full API
wb = openpyxl.load_workbook(FILE, data_only=True)
print(f"=== WORKBOOK ===")
print(f"Sheets: {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=" * 70)
    print(f"SHEET: '{sheet_name}'")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()

    # Read first 10 rows
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:
            break
        rows.append(list(row))

    if not rows:
        print("  (empty sheet)")
        continue

    # Print first 10 rows verbatim with column letters
    print("First 10 rows:")
    for i, row in enumerate(rows):
        # Truncate long cells
        truncated = []
        for c in row:
            if c is None:
                truncated.append(None)
            elif isinstance(c, str) and len(c) > 40:
                truncated.append(c[:40] + '...')
            else:
                truncated.append(c)
        print(f"  Row {i+1}: {truncated}")
    print()

    # Detect header row by keyword scanning
    keywords = ['invoice', 'date', 'party', 'amount', 'stock', 'item', 'name', 'qty', 'quantity',
                'rate', 'price', 'gst', 'total', 'bill', 'purchase', 'sale', 'inward', 'outward',
                'product', 'description', 'hsn', 'unit', 'value']
    for i, row in enumerate(rows[:6]):
        if not row:
            continue
        row_str = ' '.join(str(c).lower() if c else '' for c in row)
        matched = [kw for kw in keywords if kw in row_str]
        if len(matched) >= 3:
            print(f"*** LIKELY HEADER (row {i+1}) — matched: {matched}")
            print(f"    Headers:")
            for j, h in enumerate(row):
                if h is not None:
                    print(f"      Col {j+1} ({chr(65+j) if j < 26 else 'A'+chr(65+j-26)}): {h}")
            break
    print()

wb.close()
print("=== DONE ===")
