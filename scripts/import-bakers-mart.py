#!/usr/bin/env python3
"""
Bakers Mart DMP — Excel Import Converter
=========================================
Parses the user's business report Excel and generates a BizBook Pro
format backup Excel that can be uploaded via Backup & Restore page.

Source file: Bakers_Mart_DMP_Business_Report_April_to_June_2026.xlsx
  Sheet 1: "DMP Business Report Apr-Jun'26"  (summary — skip)
  Sheet 2: "Stock Inventory"   → 894 products  → Inventory sheet
  Sheet 3: "Sales Detail"      → 382 items     → Sales sheet
  Sheet 4: "Purchase Detail"   → 35 vouchers   → Purchases sheet

Output: Bakers_Mart_DMP_BizBook_Import.xlsx
  Sheet _README: metadata
  Sheet Inventory: 894 rows with BizBook Pro column headers
  Sheet Sales: 382 rows with BizBook Pro column headers
  Sheet Purchases: 35 rows with BizBook Pro column headers
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime, timezone
import json
import os

SOURCE = '/home/z/my-project/upload/Bakers_Mart_DMP_Business_Report_April_to_June_2026.xlsx'
OUTPUT = '/home/z/my-project/download/Bakers_Mart_DMP_BizBook_Import.xlsx'

# Generate unique IDs (cuid-like)
_id_counter = 0
def gen_id(prefix='rec'):
    global _id_counter
    _id_counter += 1
    return f'{prefix}_{_id_counter:06d}_{datetime.now().strftime("%Y%m%d%H%M%S")}'

def safe_num(val, default=0):
    """Convert to float, handling None, strings, errors."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_str(val, default=''):
    if val is None:
        return default
    return str(val).strip()

def fmt_date(dt=None):
    if dt is None:
        dt = datetime.now(timezone.utc)
    return dt.strftime('%Y-%m-%dT%H:%M:%S.%fZ')

NOW = fmt_date()

# ============================================================
# 1. Parse source Excel
# ============================================================
print(f'Loading source: {SOURCE}')
src_wb = openpyxl.load_workbook(SOURCE, data_only=True)
print(f'Source sheets: {src_wb.sheetnames}')

# --- Parse Stock Inventory ---
print('\n--- Parsing Stock Inventory ---')
ws = src_wb['Stock Inventory']
inventory_rows = []
for row in ws.iter_rows(min_row=6, values_only=True):  # data starts row 6
    if not row[1]:  # col B (Product) is empty
        continue
    product = safe_str(row[1])
    sub_group = safe_str(row[2]) if len(row) > 2 else ''
    brand = safe_str(row[3]) if len(row) > 3 else ''
    product_id = safe_str(row[4]) if len(row) > 4 else ''
    opening = safe_num(row[5]) if len(row) > 5 else 0
    inward = safe_num(row[6]) if len(row) > 6 else 0
    outward = safe_num(row[7]) if len(row) > 7 else 0
    closing = safe_num(row[8]) if len(row) > 8 else 0
    closing_value = safe_num(row[9]) if len(row) > 9 else 0

    # Calculate purchase price from closing value / closing stock
    purchase_price = round(closing_value / closing, 2) if closing > 0 else 0

    inventory_rows.append({
        'id': gen_id('inv'),
        'name': product,
        'sku': '',
        'barcode': product_id,
        'hsnCode': product_id,
        'unit': 'PCS',
        'category': sub_group,
        'brand': brand,
        'itemType': 'RAW_MATERIAL',
        'purchasePrice': purchase_price,
        'salePrice': 0,
        'mrp': 0,
        'openingStock': opening,
        'currentStock': closing,
        'minStock': 0,
        'gstRate': 0,
        'value': closing_value,
        'isDeleted': 'No',
        'deletedAt': '',
        'createdAt': NOW,
        'updatedAt': NOW,
    })
print(f'  Parsed {len(inventory_rows)} inventory items')

# --- Parse Sales Detail ---
print('\n--- Parsing Sales Detail ---')
ws = src_wb['Sales Detail']
sales_rows = []
for row in ws.iter_rows(min_row=6, values_only=True):  # data starts row 6
    if not row[2]:  # col C (Invoice Number) is empty
        continue
    date_str = safe_str(row[1])
    invoice_number = safe_str(row[2])
    product = safe_str(row[3])
    qty = safe_num(row[4])
    rate = safe_num(row[5])
    gross_amt = safe_num(row[6])
    tax = safe_num(row[7]) if len(row) > 7 else 0
    charges = safe_num(row[8]) if len(row) > 8 else 0
    discount = safe_num(row[9]) if len(row) > 9 else 0
    net_amount = safe_num(row[10]) if len(row) > 10 else gross_amt + tax + charges - discount

    # Build items JSON string
    items = [{
        'name': product,
        'qty': qty,
        'rate': round(rate, 2),
        'amount': gross_amt,
        'gstRate': 0,
        'unit': 'PCS',
        'saleItemType': 'RETAIL',
        'discount': discount,
    }]
    items_str = json.dumps(items)

    # Use 2026-04-01 as the sale date (period start — source has range text)
    sale_date = '2026-04-01T00:00:00.000Z'

    sales_rows.append({
        'id': gen_id('sale'),
        'invoiceNumber': invoice_number,
        'date': sale_date,
        'partyName': 'Cash Customer',
        'partyAddress': '',
        'partyGst': '',
        'items': items_str,
        'subtotal': gross_amt,
        'gstAmount': tax,
        'totalAmount': net_amount,
        'invoiceType': 'TAX_INVOICE',
        'invoiceStatus': 'CONFIRMED',
        'paymentStatus': 'RECEIVED',
        'amountReceived': net_amount,
        'amountPaid': net_amount,
        'notes': '',
        'einvoiceIrn': '',
        'einvoiceAckNo': '',
        'einvoiceAckDate': '',
        'einvoiceStatus': 'PENDING',
        'isDeleted': 'No',
        'deletedAt': '',
        'createdAt': NOW,
        'updatedAt': NOW,
    })
print(f'  Parsed {len(sales_rows)} sale records')

# --- Parse Purchase Detail ---
print('\n--- Parsing Purchase Detail ---')
ws = src_wb['Purchase Detail']
purchase_rows = []
for row in ws.iter_rows(min_row=6, values_only=True):  # data starts row 6
    if not row[2]:  # col C (Vch ID) is empty
        continue
    date_str = safe_str(row[1])
    vch_id = safe_str(row[2])
    vch_name = safe_str(row[3]) if len(row) > 3 else 'Purchase'
    supplier = safe_str(row[4]) if len(row) > 4 else 'Unknown Supplier'
    ref = safe_str(row[5]) if len(row) > 5 else ''
    qty = safe_num(row[6]) if len(row) > 6 else 0
    sub_total = safe_num(row[7]) if len(row) > 7 else 0
    tax = safe_num(row[8]) if len(row) > 8 else 0
    ro = safe_num(row[9]) if len(row) > 9 else 0
    bill_amt = safe_num(row[10]) if len(row) > 10 else sub_total + tax + ro

    # Parse date (e.g., "01-May-2026" → "2026-05-01T00:00:00.000Z")
    try:
        from datetime import datetime as dt
        parsed_date = dt.strptime(date_str, '%d-%b-%Y')
        purchase_date = parsed_date.strftime('%Y-%m-%dT00:00:00.000Z')
    except:
        purchase_date = '2026-04-01T00:00:00.000Z'

    # Build items JSON string
    rate = round(sub_total / qty, 2) if qty > 0 else 0
    items = [{
        'name': vch_name,
        'qty': qty,
        'rate': rate,
        'amount': sub_total,
        'gstRate': 0,
        'unit': 'PCS',
    }]
    items_str = json.dumps(items)

    purchase_rows.append({
        'id': gen_id('pur'),
        'invoiceNumber': vch_id,
        'date': purchase_date,
        'partyName': supplier,
        'partyAddress': '',
        'partyGst': '',
        'items': items_str,
        'subtotal': sub_total,
        'gstAmount': tax,
        'totalAmount': bill_amt,
        'paymentStatus': 'PAID',
        'amountPaid': bill_amt,
        'notes': f'Ref: {ref}' if ref else '',
        'einvoiceIrn': '',
        'einvoiceAckNo': '',
        'einvoiceAckDate': '',
        'einvoiceStatus': 'PENDING',
        'isDeleted': 'No',
        'deletedAt': '',
        'createdAt': NOW,
        'updatedAt': NOW,
    })
print(f'  Parsed {len(purchase_rows)} purchase records')

src_wb.close()

# ============================================================
# 2. Generate BizBook Pro format Excel backup
# ============================================================
print(f'\n--- Generating BizBook Pro backup Excel ---')
out_wb = Workbook()

# --- _README sheet ---
ws_meta = out_wb.active
ws_meta.title = '_README'
meta_data = [
    ('Key', 'Value'),
    ('Application', 'BizBook Pro'),
    ('Version', '2.0'),
    ('Export Date', NOW),
    ('Company', 'Bakers Mart - DMP'),
    ('Company ID', ''),
    ('Total Records', str(len(inventory_rows) + len(sales_rows) + len(purchase_rows))),
    ('Sheets', 'Inventory,Sales,Purchases'),
    ('Source File', 'Bakers_Mart_DMP_Business_Report_April_to_June_2026.xlsx'),
    ('Period', '01-Apr-2026 to 30-Jun-2026'),
    ('Inventory Items', str(len(inventory_rows))),
    ('Sales Records', str(len(sales_rows))),
    ('Purchase Records', str(len(purchase_rows))),
]
for row in meta_data:
    ws_meta.append(row)

# --- Inventory sheet ---
ws_inv = out_wb.create_sheet('Inventory')
inv_headers = ['id', 'name', 'sku', 'barcode', 'hsnCode', 'unit', 'category', 'brand',
               'itemType', 'purchasePrice', 'salePrice', 'mrp', 'openingStock',
               'currentStock', 'minStock', 'gstRate', 'value', 'isDeleted',
               'deletedAt', 'createdAt', 'updatedAt']
ws_inv.append(inv_headers)
for row in inventory_rows:
    ws_inv.append([row[h] for h in inv_headers])
print(f'  Inventory sheet: {len(inventory_rows)} rows')

# --- Sales sheet ---
ws_sales = out_wb.create_sheet('Sales')
sales_headers = ['id', 'invoiceNumber', 'date', 'partyName', 'partyAddress', 'partyGst',
                 'items', 'subtotal', 'gstAmount', 'totalAmount', 'invoiceType',
                 'invoiceStatus', 'paymentStatus', 'amountReceived', 'amountPaid',
                 'notes', 'einvoiceIrn', 'einvoiceAckNo', 'einvoiceAckDate',
                 'einvoiceStatus', 'isDeleted', 'deletedAt', 'createdAt', 'updatedAt']
ws_sales.append(sales_headers)
for row in sales_rows:
    ws_sales.append([row[h] for h in sales_headers])
print(f'  Sales sheet: {len(sales_rows)} rows')

# --- Purchases sheet ---
ws_pur = out_wb.create_sheet('Purchases')
pur_headers = ['id', 'invoiceNumber', 'date', 'partyName', 'partyAddress', 'partyGst',
               'items', 'subtotal', 'gstAmount', 'totalAmount', 'paymentStatus',
               'amountPaid', 'notes', 'einvoiceIrn', 'einvoiceAckNo', 'einvoiceAckDate',
               'einvoiceStatus', 'isDeleted', 'deletedAt', 'createdAt', 'updatedAt']
ws_pur.append(pur_headers)
for row in purchase_rows:
    ws_pur.append([row[h] for h in pur_headers])
print(f'  Purchases sheet: {len(purchase_rows)} rows')

# Save
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
out_wb.save(OUTPUT)
out_wb.close()

print(f'\n✅ Generated: {OUTPUT}')
print(f'   File size: {os.path.getsize(OUTPUT) / 1024:.1f} KB')
print(f'   Total records: {len(inventory_rows) + len(sales_rows) + len(purchase_rows)}')
print(f'   - Inventory: {len(inventory_rows)} items')
print(f'   - Sales: {len(sales_rows)} records')
print(f'   - Purchases: {len(purchase_rows)} records')

# ============================================================
# 3. Summary for verification
# ============================================================
total_sales = sum(r['totalAmount'] for r in sales_rows)
total_purchases = sum(r['totalAmount'] for r in purchase_rows)
total_stock_value = sum(r['value'] for r in inventory_rows)

print(f'\n--- Verification Summary ---')
print(f'   Total Sales Amount:   ₹{total_sales:,.0f}')
print(f'   Total Purchase Amount: ₹{total_purchases:,.0f}')
print(f'   Total Stock Value:     ₹{total_stock_value:,.0f}')
print(f'   Gross Margin:          ₹{total_sales - total_purchases:,.0f}')
print(f'\n   Source file totals (from summary sheet):')
print(f'   - TOTAL SALES (NET):     ₹904,292')
print(f'   - TOTAL PURCHASES:       ₹876,764')
print(f'   - CLOSING STOCK VALUE:   ₹2,485,763')
